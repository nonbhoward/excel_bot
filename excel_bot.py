from minimalog.minimal_log import MinimalLog
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from os import getcwd
from os import walk
from os.path import exists
from pathlib2 import Path
EXCEL_FORMATS = ['xlsx']  # fyi openpyxl supports four file types but only one is used here
ml = MinimalLog(__name__)


class ExcelBot:
    def __init__(self, workbook_koi, worksheet_koi):
        """
        on initialization of this class.. it does the following..
        1. create empty dictionaries
        2. set the koi, "keywords of interest", to filter data of interest
        3. set the search range, to filter ranges of interest
        4. set the data path, where excel worksheets to-be-read are stored
        5. get all excel files in the data path and load them into the class
        6. read their contents, and store those values in a dict in the class object
        7. filter the contents based on user parameters, sorting 'found results' into a dict        
        :param workbook_koi: keywords of interest that specify how workbooks will be processed
        :param worksheet_koi: keywords of interest that specify how worksheets will be processed
        """
        ml.log_event('initializing \'{}\''.format(self.__class__.__name__), event_completed=False, announce=True)
        # initialize variables
        self.search_results, self.workbooks, self.worksheets, self.worksheet_data, self.worksheet_data_of_interest = \
            dict(), dict(), dict(), dict(), dict()
        self.workbook_koi, self.worksheet_koi = workbook_koi, worksheet_koi
        self.min_col, self.max_col, self.min_row, self.max_row = self._get_default_range()
        # setup data paths and fetch file data
        self.data_path = self._get_data_path()
        self.excel_files = self._get_excel_files_from_data_path()
        self._read_worksheets_into_class()
        self._extract_worksheet_data_into_class()
        self._extract_worksheet_data_of_interest()
        ml.log_event('initializing \'{}\''.format(self.__class__.__name__), event_completed=True, announce=True)

    def search_worksheets_of_interest_and_record_cells_containing_(self, search_terms: list):
        """
        _search_worksheets_of_interest_and_record_cells_containing_ does the following..
        1. loops over worksheet data, extracting worksheet title, and worksheet data
        2. for each entry, creates a dictionary, keyed by the worksheet title
        3. for each entry, saves the parent workbook so we always know where the worksheet came from
        4. then, against each of those entries, every search term is tested is tested against every cell
        5. if a match is found.. store it in a list..
        6. ..put that list inside of a dictionary, keyed by search term..
        7. so you end up with a list of cells that contain the search term, stored in the same dictionary as above
        :param search_terms: check if a cell contains any of these terms
        :return: None, data stored in class
        """
        try:
            ml.log_event('searching worksheets of interest', event_completed=False)
            ml.log_event('record cells that contain search_terms: {}'.format(search_terms), event_completed=False)
            for worksheet_title, worksheet_data in self.worksheet_data.items():
                self.search_results[worksheet_title] = dict()
                self.search_results[worksheet_title]['parent_workbook'] = worksheet_data['parent_workbook']
                for search_term in search_terms:
                    self.search_results[worksheet_title][search_term] = list()
                    for cell_address, cell_value in worksheet_data.items():
                        if search_term in str(cell_value):
                            self.search_results[worksheet_title][search_term].append(cell_address)
                            ml.log_event('{} found in {} at {}'.format(cell_value, worksheet_title, cell_address))
            ml.log_event('record cells that contain search_terms: {}'.format(search_terms), event_completed=True)
            ml.log_event('searching worksheets of interest', event_completed=True)
        except KeyError as k_err:
            ml.log_event(k_err)

    def set_keywords_of_interest(self, workbook_keywords, worksheet_keywords):
        """
        set_keywords_of_interest does the following..
        1. allows the user to provide their own list of workbook and worksheet keywords of interest
        :param workbook_keywords: a list of workbook keywords
        :param worksheet_keywords:  a list of worksheet keywords
        :return:  None, data stored in class
        """
        try:
            ml.log_event('set workbook_keywords: \'{}\' and worksheet keywords: \'{}\''.format(
                workbook_keywords, worksheet_keywords))
            self.workbook_koi, self.worksheet_koi = workbook_keywords, worksheet_keywords
        except ValueError as v_err:
            ml.log_event(v_err)

    def set_search_area(self, search_area):
        """
        set_search_area does the following..
        1. allows the user to set a custom search area for all spreadsheets, smaller is faster
        :param search_area: a tuple containing..
                                ..min_col: far left,
                                ..max_col: far right,
                                ..min_row: top,
                                ..max_row: bottom
        :return:
        """
        min_col, max_col, min_row, max_row = search_area
        ml.log_event('search area set..\n min_col: \'{}\'\n max_col: \'{}\'\n min_row: \'{}\'\n max_row: \'{}\''.format(
            min_col, max_col, min_row, max_row))
        self.min_col = min_col
        self.max_col = max_col
        self.min_row = min_row
        self.max_row = max_row

    @staticmethod
    def write_file_to_disk(output_filename):
        """
        perform_write_operations does the following..
        1. check if the output file exists
        2. if it exists, do nothing
        3. if it does not exist..
        4. initialize a new workbook
        5. save the new workbook using the output_filename
        :param output_filename: output filename to write
        :return: None
        """
        ml.log_event('performing write operations on file \'{}\''.format(output_filename), event_completed=False)
        try:
            if not exists(output_filename):
                new_workbook = Workbook()
                new_workbook.save(output_filename)
                ml.log_event('performing write operations on file \'{}\''.format(output_filename), event_completed=True)
                return
            ml.log_event('warning, file already exists')
            ml.log_event('performing write operations on file \'{}\''.format(output_filename), event_completed=True)
        except OSError as o_err:
            ml.log_event(o_err)

    def _extract_data_from_worksheet(self, workbook_path, a_worksheet):
        """
        _extract_data_from_worksheet does the following..
        1. creates a worksheet_data dictionary
        2. saves the parent workbook so we can always find out what workbook the worksheet came from
        3. loops over the entire search area
        4. saves all values, if the cell is empty, saves nothing
        :param workbook_path: the path to the workbook to be read
        :param a_worksheet: the worksheet name to be read
        :return: None
        """
        ml.log_event('extracting data from worksheet: \'{}\' at workbook_path: \'{}\''.format(
            a_worksheet, workbook_path), event_completed=False)
        try:
            self.worksheet_data[a_worksheet.title] = dict()
            self.worksheet_data[a_worksheet.title]['parent_workbook'] = str(workbook_path)
            for column in range(1, self.max_col + 1):
                for row in range(1, self.max_row + 1):
                    cell = get_column_letter(column) + str(row)
                    value = a_worksheet[cell].value
                    if value is not None:
                        ml.log_event('value: {} found at cell: {}'.format(value, cell))
                        self.worksheet_data[a_worksheet.title][cell] = value
            ml.log_event('extracting data from worksheet: {} at workbook_path: {}'.format(
                a_worksheet, workbook_path), event_completed=True)
        except KeyError as k_err:
            ml.log_event(k_err)

    def _extract_worksheet_data_into_class(self):
        """
        _extract_worksheet_data_into_class does the following..
        1. for all excel worksheets, extracts their path and worksheets
        2. for each worksheets object, extract a worksheet
        3. for each worksheet
        4. extract data according to _extract_data_from_worksheet() docstring
        :return: bool, successful if True
        """
        ml.log_event('extract worksheet data into class', event_completed=False)
        try:
            read_data_success = list()
            for workbook_path, worksheets_dict in self.worksheets.items():
                for worksheet_title in worksheets_dict.keys():
                    a_worksheet = worksheets_dict[worksheet_title]
                    self._extract_data_from_worksheet(workbook_path, a_worksheet)
            if all(read_data_success):
                ml.log_event('extract worksheet data into class', event_completed=True)
                return True
            ml.log_event('extract worksheet data into class, failure encountered', event_completed=True)
            return False
        except OSError as o_err:
            ml.log_event(o_err)

    @staticmethod
    def _get_data_directory_name() -> str:
        """
        _get_data_directory_name does the following..
        1. returns the name of a folder where data files are kept
        :return: a string representing a directory name that should exist and contain excel files
        """
        ml.log_event('get data directory name')
        return 'data_src'

    def _get_data_path(self) -> Path:
        """
        _get_data_path does the following..
        1. builds a path object from project directory and data directory
        2. returns it
        :return: a path object with the complete path pointing toward where excel files are stored
        """
        ml.log_event('get data path', event_completed=False)
        try:
            data_path = Path(self._get_project_directory(), self._get_data_directory_name())
            ml.log_event('get data path', event_completed=True)
            return data_path
        except RuntimeError as r_err:
            ml.log_event(r_err)

    @staticmethod
    def _get_default_range() -> tuple:
        """
        _get_default_range does the following..
        1. provides a default range in case the user forgets
        :return: a tuple containing four integers representing a search area on an excel file
        """
        ml.log_event('get default range')
        return 1, 100, 1, 100

    def _get_data_path_files(self) -> list:
        """
        _get_data_path_files does the following..
        1. searches through the data path for every directory and file
        2. converts each file into a path that can be used to directly reference that file
        3. saves them all into a list
        4. returns the list
        :return: a list of all files, with complete paths, in the data directory
        """
        ml.log_event('get files from data path', event_completed=False)
        try:
            all_files = list()
            for root, dirs, files in walk(self.data_path):
                for file in files:
                    all_files.append(Path(root, file))
            ml.log_event('get files from data path', event_completed=True)
            return all_files
        except FileNotFoundError as f_err:
            ml.log_event(f_err)

    def _get_excel_files_from_data_path(self) -> list:
        """
        _get_excel_files_from_data_path does the following..
        1. gets a list of all files in the data path
        2. loops through every single one of them
        3. if it finds an excel file it adds it to the list of excel files
        4. return a list of excel files that are in the data path
        :return: a list of excel files that are in the data path
        """
        ml.log_event('get excel files from data path', event_completed=False)
        try:
            excel_files = list()
            data_path_files = self._get_data_path_files()
            for file in data_path_files:
                if _is_excel_file(file):
                    excel_files.append(file)
            ml.log_event('get excel files from data path', event_completed=True)
            return excel_files
        except FileNotFoundError as f_err:
            ml.log_event(f_err)

    @staticmethod
    def _get_project_directory() -> Path:
        """
        _get_project_directory does the following..
        1. uses os.getcwd() to get a string to the current project directory
        2. converts it into a Path because Paths are easier to work with
        3. returns the Path to the project
        :return: a Path object to the project directory
        """
        ml.log_event('get project directory')
        try:
            return Path(getcwd())
        except RuntimeError as r_err:
            ml.log_event(r_err)

    def _extract_worksheet_data_of_interest(self):
        """
        _extract_worksheet_data_of_interest does the following..
        1. for each worksheet containing data..
        2. for each workbook keyword..
        3. checks to see if the worksheet's parent workbook's name matches any of the workbook keywords
        4. if it doesn't, do nothing
        5. if it does, check to see if any of that workbook's worksheets match any of the worksheet keywords
        6. if none of them do, do nothing
        7. if the worksheet matches the keyword (workbook already matched)
        8. save the worksheet data to a special dictionary indicating it is of special interest
        9. once the dictionary is built, save it to the class at self.worksheet_data_of_interest
        :return: None
        """
        ml.log_event('extract worksheet data of interest', event_completed=False)
        try:
            worksheet_data_of_interest = dict()
            for worksheet_title, worksheet_data in self.worksheet_data.items():
                for workbook_keywords in self.workbook_koi:
                    if workbook_keywords in worksheet_data['parent_workbook']:
                        for worksheet_keyword in self.worksheet_koi:
                            if worksheet_keyword in worksheet_title:
                                ml.log_event('workbook of interest found: {} worksheet of interest found: {}'.format(
                                    worksheet_data['parent_workbook'], worksheet_title))
                                worksheet_data_of_interest[worksheet_title] = worksheet_data
            self.worksheet_data_of_interest = worksheet_data_of_interest
            ml.log_event('extract worksheet data of interest', event_completed=True)
        except RuntimeError as r_err:
            ml.log_event(r_err)

    def _read_worksheets_into_class(self):
        """
        _read_worksheets_into_class does the following..
        1. for each excel file in the data directory..
        2. save the excel objects into the class
        :return: bool, True if *all* reads were successful, False if any failed
        """
        ml.log_event('read worksheets into class', event_completed=False)
        try:
            excel_file_read_success = list()
            for excel_file in self.excel_files:
                if self._worksheets_saved_to_class(excel_file):
                    ml.log_event('worksheets from {} saved to class'.format(excel_file))
                    excel_file_read_success.append(True)
            if all(excel_file_read_success):
                ml.log_event('read worksheets into class', event_completed=True)
                return True
            ml.log_event('read worksheets into class, a failure was encountered', event_completed=True)
            return False
            pass
        except RuntimeError as r_err:
            ml.log_event(r_err)

    def _worksheets_saved_to_class(self, excel_file):
        """
        _worksheets_save_to_class does the following..
        1. takes an excel file as an argument..
        2. create a workbook object, using the existing excel file
        3. create a worksheet dictionary, keyed by the path of the parent workbook
        4. for each worksheet in a workbook..
        5. save each worksheet object to the class, the worksheet is stored in a dictionary..
        ..which is stored in another dictionary. the data structure looks like :
        self.worksheets[excel_file_full_path][worksheet_title] = worksheet_data
        :param excel_file: an excel file to process into a workbook, all data is saved into the class
        :return: bool, if the
        """
        ml.log_event('save worksheets from {} to class'.format(excel_file), event_completed=False)
        try:
            self.workbooks[excel_file] = load_workbook(excel_file)
            self.worksheets[excel_file] = dict()
            for misc_worksheet in self.workbooks[excel_file]:
                worksheet_title = misc_worksheet.title
                self.worksheets[excel_file][worksheet_title] = misc_worksheet
            worksheet_count = len(self.worksheets[excel_file])
            if worksheet_count > 0:
                # TODO this is a bug
                # TODO fyi this will be true even if a read fails, since as long as the first read is..
                # TODO ..successful then len(self.worksheets[]) will always be non-zero
                ml.log_event('save worksheets from {} to class'.format(excel_file), event_completed=False)
                return True
            ml.log_event('failure to save worksheets from {} to class'.format(excel_file))
            return False
        except KeyError as k_err:
            ml.log_event(k_err)


def _is_excel_file(filename: Path) -> bool:
    """
    _is_excel_file does the following..
    1. takes a path object that should be an excel file
    2. splits it into pieces, using a period punctuation mark as the delimiter
    3. checks if the file is excel by looking at the final element in the array, the extension
    Note: openpyxl supports four excel extensions, add them to the list EXCEL_FORMATS if needed
    :param filename: the path to the excel file
    :return: boolean, True is file ends in a user-accepted excel extension
    """
    ml.log_event('check if {} is excel file'.format(filename), event_completed=False)
    try:
        name_parts = str(filename).split('.')
        if name_parts[-1] in EXCEL_FORMATS:
            ml.log_event('check if {} is excel file'.format(filename), event_completed=True)
            return True
        ml.log_event('{} failed check'.format(filename))
        return False
    except IndexError as i_err:
        ml.log_event(i_err)
