from directory_utils.directory_utils import *
from minimalog.minimal_log import MinimalLog
from openpyxl import load_workbook, Workbook, worksheet
from pickle import dumps, load
from string import ascii_letters, ascii_uppercase, digits
from time import sleep
ml = MinimalLog()
SUCCESSFUL = True


class ExcelTasker:
    def __init__(self, read=False, write=False, debug=False, fetch_downloads=False):
        """
        :param read: bool, populate data for read operations
        :param write: bool, create workbook objects for write operations
        :param fetch_downloads: bool, try to copy newly discovered files from downloads to data path
        :param debug: bool, debug function will run after initialization
        """
        ml.log_event(event='init ExcelTasker with read = {} and write = {}'.format(read, write),
                     event_completed=False)
        self.open_workbooks, self.workbooks_to_create = dict(), dict()
        self.path, self.data_dir, self.downloads_dir = get_path_at_cwd(), get_data_path(), get_os_downloads_path()
        if write:
            ml.log_event(event='write mode init', event_completed=False)
            self.workbook_status, self.created_workbooks = dict(), dict()
            self.workbooks_to_create = self.append_xlsx_extension_to_filenames(self.get_filenames())
            self.create_workbooks_in_queue()
            ml.log_event(event='write mode init', event_completed=True)
        self.files_in_data_path = get_all_files_in(self.data_dir)
        if fetch_downloads:
            pass  # TODO fetch files from downloads & move to data dir
        self.workbook_names = filter_files_by_ext(files=self.files_in_data_path, valid_extensions=EXCEL_EXTS)
        if read:
            ml.log_event('read mode init', event_completed=False)
            self.open_workbooks = self.open_excel_files_in_data_dir()
            self.extract_data_range_from_open_worksheets()
            ml.log_event('read mode init', event_completed=True)
        ml.log_event(event='init ExcelTasker', event_completed=True)
        if debug:
            ml.log_event('debug routine', event_completed=False)
            self.__debug()
            ml.log_event('debug routine', event_completed=True)

    @staticmethod
    def append_xlsx_extension_to_filenames(filenames_to_build: list) -> dict:
        """
        splits filename string on '.' delimiter, ideally finds zero or two, appends '.xlsx' to first
        :param filenames_to_build: list_of_filenames
        :return: filename.xlsx
        """
        xlsx_ext = '.xlsx'
        filenames_with_ext = dict()
        try:
            for filename in filenames_to_build:
                if '.' not in filename:
                    filenames_with_ext[filename] = filename + xlsx_ext
                else:
                    string_parts = filename.split('.')
                    if len(string_parts) > 2:
                        ml.log_event('potential error with filename {}'.format(filename), level=ml.WARN)
                    filenames_with_ext[filename] = string_parts[0] + xlsx_ext
            return filenames_with_ext
        except OSError as o_err:
            ml.log_exception(o_err)

    def create_workbooks_in_queue(self):
        """
        :return: create the workbook objects queued from self
        """
        try:
            for k, file_val in self.workbooks_to_create.items():
                full_file_path = build_full_path_to_filename(self.data_dir, file_val)
                self.workbook_status[full_file_path] = dict()
                if exists(full_file_path):
                    ml.log_event('warning, file {} already exists'.format(full_file_path))
                    self.workbook_status[full_file_path]['newly_created'] = False
                    self.workbook_status[full_file_path]['exists'] = True
                    continue
                self.instantiate_and_create_workbook_at(full_file_path)
                if exists(full_file_path):
                    self.workbook_status[full_file_path]['newly_created'] = True
                    self.workbook_status[full_file_path]['exists'] = True
                else:
                    self.workbook_status[full_file_path]['newly_created'] = False
                    self.workbook_status[full_file_path]['exists'] = False
        except KeyError as k_err:
            ml.log_exception(k_err)

    def create_worksheet_name_in_workbook(self, ws_name: str, workbook: Workbook) -> bool:
        pass

    def extract_data_range_from_open_worksheets(self, top_left='a1', bottom_right='z999'):
        """
        abstract : for a_worksheet in worksheets -> init dict -> build data -> save data -> continue
        :param top_left:
        :param bottom_right:
        :return:
        """
        if self.open_workbooks is None:
            raise OSError('There are no open workbooks')
        for key in self.open_workbooks.keys():
            for a_worksheet in self.open_workbooks[key]['workbook'].worksheets:
                worksheet_title = a_worksheet.title
                cell_data = self._build_and_store_cell_data(active_workbook=self.open_workbooks[key]['workbook'],
                                                            top_left=top_left, bottom_right=bottom_right)
                self.open_workbooks[key][worksheet_title] = cell_data

    @staticmethod
    def get_filenames() -> list:
        # TODO debug function, will be replaced with something more substantial after testing
        return ['z1', 'z2', 'z3', 'z4', 'z5', 'z6', 'z7', 'z8', 'z9']

    def instantiate_and_create_workbook_at(self, path: Path) -> bool:
        """
        TODO : note that each file takes > 1s to create, asyncio opportunity here?
        :param path: path where we will save the instantiated file
        :return: boolean, success or failure
        """
        try:
            new_wb = Workbook()
            new_wb.save(path)
            full_path = str(path.resolve())
            self.created_workbooks[full_path] = dict()
            self.created_workbooks[full_path]['workbook'] = new_wb
        except OSError as o_err:
            ml.log_exception(o_err)
        if exists(path):
            return True
        return False

    def get_active_worksheet_in_active_workbook(self, worksheet_key: str) -> worksheet:
        pass

    def get_all_worksheets_from_all_open_workbooks(self) -> list:
        """
        :return: all worksheets from all open workbooks
        """
        all_worksheets = list()
        try:
            for workbook_name, workbook in self.open_workbooks.items():
                worksheets_from_this_workbook = self.get_all_worksheets_from_workbook(workbook_name)
                for a_worksheet in worksheets_from_this_workbook:
                    all_worksheets.append(a_worksheet)
            return all_worksheets
        except KeyError as k_err:
            ml.log_exception(k_err)

    def get_all_worksheets_from_workbook(self, workbook_name: str) -> list():
        """
        :param workbook_name: string that accesses workbook from workbook dictionary
        :return: list of all worksheets in the given workbook
        """
        worksheet_names = list()
        try:
            for _worksheet in self.open_workbooks[workbook_name]['workbook'].worksheets:
                worksheet_names.append(_worksheet)
            return worksheet_names
        except OSError as o_err:
            ml.log_exception(o_err)

    def get_value_at_cell(self) -> str:
        pass

    def get_worksheet_from_workbook(self, worksheet_name: str) -> worksheet:
        pass

    def open_excel_files_in_data_dir(self) -> dict:
        """
        :return: dictionary of excel workbooks, keyed by full path name
        """
        ml.log_event('open excel files {}'.format(self.workbook_names), event_completed=False)
        workbook_metadata, open_workbooks = list(), dict()
        for excel_file in self.workbook_names:
            workbook_metadata.append(self.open_excel_workbook(excel_file))
            for metadata in workbook_metadata:
                for data in metadata:
                    open_workbooks[data[0]] = dict()
                    open_workbooks[data[0]]['workbook'] = data[1]
        ml.log_event('open excel files {}'.format(self.workbook_names), event_completed=True)
        return open_workbooks

    def open_excel_workbook(self, excel_file: str) -> Workbook:
        """
        :param excel_file: string representing a filename without full path
        :return: workbook filename, workbook data
        """
        try:
            ml.log_event('open excel file {}'.format(excel_file), event_completed=False)
            excel_file = str(Path(str(self.data_dir), excel_file))
            wb = load_workbook(excel_file)
            workbook_data = [excel_file, wb]
            ml.log_event('open excel file {}'.format(excel_file), event_completed=True)
            yield workbook_data
        except OSError as o_err:
            ml.log_exception(o_err)

    def read_value_from_worksheet(self, workbook: Workbook, worksheet: worksheet, col: str, row: str) -> str:
        """
        :param workbook: the workbook to operate in
        :param worksheet: the worksheet to operate in
        :param col: the column
        :param row: the row
        :return: the string inside of the cell
        """
        try:
            cell = self._sanitize_col(col) + self._sanitize_row(row)
            cell_value = workbook[worksheet][cell].value
            return cell_value
        except OSError as o_err:
            ml.log_exception(o_err)

    def search_active_worksheet_for_cell_value(self) -> dict():
        pass

    def set_active_worksheet(self, active_workbook: Workbook, worksheet_title: str) -> bool:
        if worksheet_title not in active_workbook['workbook'].sheetnames:
            raise OSError('worksheet title not found in workbook {}'.format(active_workbook))
        try:
            self.active_worksheet = active_workbook[worksheet_title]
        except OSError as o_err:
            ml.log_exception(o_err)

    def _build_and_store_cell_data(self, active_workbook: Workbook, top_left: str, bottom_right: str) -> dict:
        """
        :param active_workbook: workbook containing data
        :param top_left: top left cell for data range
        :param bottom_right: bottom right cell for data range
        :return: bool, success
        """
        try:
            cell_dict = self._generate_cells(top_left_cell=top_left, bottom_right_cell=bottom_right)
            for cell_address in cell_dict.keys():
                cell_dict[cell_address] = active_workbook.active[cell_address].value
            return self._purge_none_from_dict(cell_dict)
        except KeyError as k_err:
            ml.log_exception(k_err)

    def _extract_column_data(self, cell_a: str, cell_b: str) -> tuple:
        cell_a_column_data, cell_b_column_data = list(), list()
        cell_a, cell_b = self._sanitize_col(cell_a), self._sanitize_col(cell_b)
        for character in cell_a:
            if character in ascii_uppercase:
                cell_a_column_data.append(character)
        for character in cell_b:
            if character in ascii_uppercase:
                cell_b_column_data.append(character)
        return ''.join(cell_a_column_data), ''.join(cell_b_column_data)

    def _extract_row_data(self, cell_a: str, cell_b: str) -> tuple:
        cell_a_row_data, cell_b_row_data = list(), list()
        cell_a, cell_b = self._sanitize_row(cell_a), self._sanitize_row(cell_b)
        for character in cell_a:
            if character in digits:
                cell_a_row_data.append(character)
        for character in cell_b:
            if character in digits:
                cell_b_row_data.append(character)
        return ''.join(cell_a_row_data), ''.join(cell_b_row_data)

    def _generate_cells(self, top_left_cell: str, bottom_right_cell: str) -> dict:
        cell_dict = dict()
        cells = self._extract_column_data(top_left_cell, bottom_right_cell)
        rows = self._extract_row_data(top_left_cell, bottom_right_cell)
        columns = self._generate_columns(min_col=cells[0], max_col=cells[1])
        rows = self._generate_rows(min_row=rows[0], max_row=rows[1])
        for column in columns:
            for row in rows:
                cell_dict[column+row] = ''
        return cell_dict

    def _generate_columns(self, min_col: str, max_col: str) -> list:
        """
        :param min_col:
        :param max_col:
        :return:
        """
        ml.log_event('generating columns from {} and {}'.format(min_col, max_col))
        generated_cols, min_col, max_col, record = list(), self._sanitize_col(min_col), \
                                                   self._sanitize_col(max_col), False
        column_range, columns, record = list(), self._generate_column_sample_oversized(), False
        for column in columns:
            if min_col == column:
                record = True
            if record:
                column_range.append(column)
            if max_col == column:
                record = False
                break
        return column_range

    @staticmethod
    def _generate_column_sample_oversized() -> list:
        """
        TODO work in progress, lazy
        :return: a lot of rows
        """
        columns, ord_offset, chr_mod = list(), 65, 26
        for column_lead in range(99):
            if len(columns) < pow(26, 1):
                columns.append(chr(column_lead % chr_mod + ord_offset))
            elif len(columns) < pow(26, 2) + 26:
                for letter_upper in ascii_uppercase:
                    columns.append(chr(column_lead % chr_mod + ord_offset) + letter_upper)
            elif len(columns) < pow(26, 3) + 26:
                for letter_upper_one in ascii_uppercase:
                    for letter_upper_two in ascii_uppercase:
                        columns.append(chr(column_lead % chr_mod + ord_offset) + letter_upper_one + letter_upper_two)
        return columns

    @staticmethod
    def _generate_rows(min_row: int, max_row: int) -> list:
        generated_rows = list()
        for row in range(int(min_row), int(max_row) + 1):
            generated_rows.append(str(row))
        return generated_rows

    def _get_active_workbook(self, wb_key_substring='') -> Workbook:
        """
        :param wb_key_substring: the full path to the file, including extension, is the workbook key
        :return: a workbook instance
        """
        try:
            if wb_key_substring == '':
                for workbook in self.open_workbooks:
                    return workbook  # if no key provided, return the first workbook found
            for workbook_key in self.open_workbooks.keys():
                # find a workbook_key that contains wb_key_substring and return the associated workbook
                if wb_key_substring in workbook_key:
                    return self.open_workbooks[workbook_key]
            return None  # TODO what else could be done here?
        except KeyError as k_err:
            ml.log_exception(k_err)

    def _get_active_worksheet(self, ws_key_substring='') -> worksheet:
        try:
            if ws_key_substring == '':
                for _worksheet in self.active_workbook['workbook'].worksheets:
                    return _worksheet  # if no key provided, return the first _worksheet in the active workbook
            for _worksheet in self.active_workbook['workbook'].worksheets:
                # find a workbook_key that contains wb_key_substring and return the associated workbook
                if ws_key_substring in _worksheet.title:
                    return _worksheet
        except KeyError as k_err:
            ml.log_exception(k_err)

    def _purge_none_from_dict(self, none_dict: dict) -> dict:
        """
        :param none_dict: dict potentially containing None values
        :return: dict containing zero None values
        """
        keys_with_none_values = list()
        try:
            for key in none_dict:
                if none_dict[key] is None:
                    keys_with_none_values.append(key)
            for key in keys_with_none_values:
                none_dict.pop(key)
            return none_dict
        except KeyError as k_err:
            ml.log_exception(k_err)

    @staticmethod
    def _sanitize_col(col: str) -> str:
        """
        :param col: a letter, or series of letters, representing a column
        :return: col -> remove non-letters -> upper() -> return
        """
        sanitized_col = list()
        try:
            for letter in col:
                if letter in ascii_letters:
                    sanitized_col.append(letter.upper())
            return ''.join(sanitized_col)
        except IndexError as i_err:
            ml.log_exception(i_err)

    @staticmethod
    def _sanitize_row(row: str) -> str:
        """
        :param row: a number, or series of numbers, representing a row
        :return: row -> remove non-digits -> return
        """
        sanitized_row = list()
        try:
            for digit in row:
                if digit in digits:
                    sanitized_row.append(digit)
            return ''.join(sanitized_row)
        except IndexError as i_err:
            ml.log_exception(i_err)

    def _write_value_at_cell(self, col: str, row: str, value: str) -> bool:
        """
        :param col: a letter, or series of letters, representing a column
        :param row: a number, or series of numbers, representing a row
        :param value: a string value to write to the cell
        :return: bool, success or failure
        """
        cell = self._sanitize_col(col) + self._sanitize_row(row)

    def __debug(self):
        pass


if __name__ == '__main__':
    from data_src.CONSTANTS import EXCEL_EXTS
    ml.log_event('execute ExcelTask', event_completed=False, announce=True)
    et_read = ExcelTasker(debug=True, read=True)
    ml.log_event('close ExcelTask', event_completed=True, announce=True)
    pass
else:
    from .data_src.CONSTANTS import EXCEL_EXTS
    print('importing {}'.format(__name__))
    ml.log_event('importing {}'.format(__name__))
